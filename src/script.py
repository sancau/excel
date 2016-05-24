
import re
import os, sys, traceback
import openpyxl

from openpyxl.compat import range
from openpyxl.cell import get_column_letter

from config import *


"""
LOGIC
""" 
def remove_spaces(value):
        """
        Returns given string with all the spaces removed.
        """
        return re.sub('[\s+]', '', str(value))       
            
def get_files(dir_path=DEFAULT_DIR_PATH):
    """
    Returns list of excel file paths in
    a given directory
    """
    result = []
    for path, subdirs, files in os.walk(dir_path):
        for file in files:
            filename, file_extension = os.path.splitext(file)
            if file_extension in ['.xls', '.xlsx']:
                file_path = os.path.join(path, file)
                result.append(file_path)   
    return result

def getValueWithMergeLookup(sheet, cell):
    idx = cell.coordinate
    for range_ in sheet.merged_cell_ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(range_))
        for row in merged_cells:
            if idx in row:
                # If this is a merged cell,
                # return  the first cell of the merge range
                return sheet.cell(merged_cells[0][0]).value
    return sheet.cell(idx).value

def pre_process(rows):
    """
    Filters and preprocess data to make it ready for the merge.
    """
    filtered_rows = [
        row for row in rows if row[NAME] is not None and row[INDEX] is not None
    ]
    payloaded = []
    for row in filtered_rows:
        payloaded_row = []
        if type(row[0]) == int:
            for idx in PAYLOAD_DATA_INDEXES:
                payloaded_row.append(row[idx])
            payloaded.append(payloaded_row)
    #handle repeat sybmols
    for row_index, row in enumerate(payloaded):
        for value_index, value in enumerate(row):
            if value in REPEAT_SYMBOLS:
                row[value_index] = payloaded[row_index - 1][value_index] 
            elif value is None:
                row[value_index] = UNDEFINED_SYMBOL
            row[value_index] = str(row[value_index])
    return payloaded 


class ExtraParamsObject:
    def __init__(self, size, amount):
        self.size = size
        self.amount = amount

        
class OutputRow:
    def form_extra_params(self, size, amount):
        params = []
        obj = ExtraParamsObject(size, amount)
        params.append(obj)
        return params
    
    def get_verbose_size(self, size):
        size_char = self.primary_size[0]
        if size_char in SIZE_DICT:
            return ' '.join([
                    SIZE_DICT[size_char]['verbose'],
                    self.primary_size[1:],
                    SIZE_DICT[size_char]['units']])
        else:
            return 'get_verbose_size returned null'
            
    
    def __init__(self, data):
        data_index = 0
        data_name = 1
        data_size = 2
        data_amount = 3
        data_material = 4
        data_units = 5
        data_material_amount = 6
        data_standart = 7
        data_comment = 8
        
        # service props
        self.material = data[data_material]
        self.primary_size = remove_spaces(str(data[data_size])).split('×')[0] 
        self.name = data[data_name]
        
        self.name_material = '%s %s %s' % (data[data_name], data[data_material], self.get_verbose_size(data[data_size]))
        self.nomen = '-'
        self.extra_params = self.form_extra_params(
            data[data_size],
            data[data_amount]
        )  
        self.nomen_number = '-'
        self.code = '-'
        self.category = '-'
        self.standart = data[data_standart]
        self.units = data[data_units]
        self.amount = data[data_material_amount]


def merge_row(output, row):
    """
    Merges given row with existing output if need.
    """
    def is_match(existed, new):
        return existed.standart == new.standart and \
                existed.material == new.material and \
                existed.primary_size == new.primary_size               
    
    new = OutputRow(row)
    if not output:
        output.append(new)
    else:
        match = [existed for existed in output if is_match(existed, new)]
        if not match: output.append(new)
        elif len(match) > 1: 
            print('Invalid parsing algorythm')
        else:
            match = match[0]
            merge_target = output[output.index(match)]
            if not merge_target.name_material == new.name_material:
                merge_target.name_material = ', '.join([new.name, merge_target.name_material])
            print("match on primary size detected")
            # check if size is equal
            # if so merge like [s-a1+a2, ...]     
            extra_param_equal = [item for item in match.extra_params \
                      if item.size == new.extra_params[0].size]
            if extra_param_equal:
                equal = extra_param_equal[0]
                print('size equal match detected')
                for_edit = match.extra_params[match.extra_params.index(equal)]
                # кол-во заготовок складывается
                for_edit.amount = str(int(new.extra_params[0].amount) + int(for_edit.amount))
                return output
            # else merge in list [s-a; s-a...]
            merge_target.extra_params += new.extra_params             
    return output

def merge(rows):
    output = []
    print('Processing output...')
    print(' ')
    rows = pre_process(rows)
    for row in rows:
        output = merge_row(output, row)
    to_file_format = []
    for item in output:
        amount_units = ' '.join([item.amount, item.units])
        obj = [
            item.name_material,
            item.nomen,
            '; '.join(['; '.join([(' - '.join([extra_param.size, ' '.join([extra_param.amount, AMOUNT_UNITS_VERBOSE])])) \
                       for extra_param in item.extra_params]), amount_units, item.standart]),
            item.nomen_number,
            item.code,
            item.category,
            item.standart,
            item.units,
            item.amount
        ]
        
        for index, value in enumerate(obj):
            if value == '-':
                obj[index] = ''

        to_file_format.append(obj)   
        
    return to_file_format

def build_results_file(rows, result_file_path):
    """
    Build an excel file based on results dict and 
    a given path.
    """
    wb = openpyxl.load_workbook('template.xlsx')
    dest_filename = os.path.join(result_file_path, DEFAULT_RESULT_FILE_NAME)
    ws = wb.active   
    for row in rows:
        for value_index, value in enumerate(row):
            row[value_index] = str(value).encode('utf-8')
        ws.append(row)   
    wb.save(filename = dest_filename)
   
def process_files(dir_path=DEFAULT_DIR_PATH, result_file_path=DEFAULT_RESULT_FILE_PATH):
    """
    Application level logic.
    """
    try:
        files = get_files(dir_path)
        rows_to_process = []
        if files:
            for file in files:
                workbook = openpyxl.load_workbook(filename=file)   
                for sheet in workbook:
                    for row in sheet:
                        # add rows by certain condition
                        row_index = (lambda x: x[0].row)(row)
                        if (sheet is not workbook[FISRT_LIST] \
                         and row_index >= OTHER_LISTS_FIRST_DATA_ROW) \
                            or row_index >= FISRT_LIST_FIRST_DATA_ROW:                           
                            merged_cells_awared_row = []
                            for cell in row:
                                value = getValueWithMergeLookup(sheet, cell)
                                merged_cells_awared_row.append(value)                                                       
                            rows_to_process.append(merged_cells_awared_row)                            
            result = merge(rows_to_process) 
            print(len(result))            
            build_results_file(result, result_file_path)            
            print(' ')
            print('Success')
        else:
            print('No files to process')    
    except Exception as ex:
        print('Error while processing')
        print(ex)
        traceback.print_exc()
